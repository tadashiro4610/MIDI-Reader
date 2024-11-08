from time import sleep
import mido
import openpyxl as op
import math as m
import numpy as np
import copy
import traceback
from functools import wraps
from tkinter import messagebox

class midiReader:
    def __init__(self,dir_midi,dir_save,bpm,delay):
        try:
            wb=op.load_workbook(dir_save)
            ws=wb.worksheets[0]
            
            def show_on_error(function):
                "関数実行中に例外が発生したら、showerrorするデコレータ"
                @wraps(function)
                def show_error(*args,**kwargs):
                    try:
                        function(*args,**kwargs)
                    except Exception as e:
                        print(traceback.format_exc())
                        title = e.__class__.__name__
                        message = traceback.format_exc()
                        messagebox.showerror(f"{title}",
                                            f"{message}")
                return show_error
            
            def getNearestValue(list, num):
                """
                概要: リストからある値に最も近い値を返却する関数
                @param list: データ配列
                @param num: 対象値
                @return 対象値に最も近い値
                """
                # リスト要素と対象値の差分を計算し最小値のインデックスを取得
                idx = np.abs(np.asarray(list) - num).argmin()
                return list[idx]
            
            @show_on_error
            def write_list_2d(sheet, l_2d, start_row, start_col):
                for i, value in enumerate(l_2d):
                    ws.cell(row=start_row + i, column=start_col, value=value)
                    
            def remove_duplicates(lst):
                return list(set(lst))

            # MIDIファイルをロード
            midi_file = mido.MidiFile(dir_midi)

            # 初期テンポ（マイクロ秒/拍）。デフォルトは500,000（120 BPM）
            tempo = mido.bpm2tempo(float(bpm))
            # print(tempo)
            ticks_per_beat = midi_file.ticks_per_beat

            # 音名リスト
            note_names = ['C', 'C♯', 'D', 'D♯', 'E', 'F', 'F♯', 'G', 'G♯', 'A', 'A♯', 'B']
            
            # バーの時間
            bar_sec=list(np.arange(0,60,60/float(bpm)/2))
            # print(bar_sec)
            # ティックを秒に変換する関数
            def ticks_to_seconds(ticks, tempo, ticks_per_beat):
                return (ticks / ticks_per_beat) * (tempo / 1_000_000)

            # 各ノートの開始時間を保存する辞書（キー: 音名, 値: 開始時間のリスト）
            note_start_times = {}

            # 時間をトラックする変数
            current_time_ticks = 0

            # 各トラックのノートオンイベントを取得
            for i, track in enumerate(midi_file.tracks):
                for msg in track:
                    # print(msg)
                    current_time_ticks += msg.time  # 現在のティック数を更新
                    # print(current_time_ticks)
                    # print(msg.time)
                

                    # ノートオンイベントの処理
                    if msg.type =="note_on" and msg.velocity == 0:
                        pass
                    elif msg.type == 'note_on' and msg.velocity > 0:
                        current_time_seconds = mido.tick2second(current_time_ticks,ticks_per_beat,tempo)
                        
                        # 音名とオクターブを計算
                        note_number = msg.note
                        note_name = note_names[note_number % 12]
                        octave = (note_number // 12) - 1  # MIDIオクターブの計算

                        full_note_name = f'{note_name}{octave}'

                        # 音名ごとに配列に開始時間を追加
                        if full_note_name not in note_start_times:
                            note_start_times[full_note_name] = []  # 初回の場合はリストを作成
                        note_start_times[full_note_name].append(current_time_seconds)

            # # 結果を表示
            # for note, times in note_start_times.items():
            #     print(f'Note {note} starts at times: {times}\n')
            
            # ノードをまとめる
            Kick=[]
            Crash=[]
            Hihat=[]
            Snare=[]
            
            if "C2" in note_start_times:
                Kick.extend(note_start_times["C2"])
            if "G3" in note_start_times:
                Crash.extend(note_start_times["G3"])
            if "A♯2" in note_start_times:
                Hihat.extend(note_start_times["A♯2"])
            if "D2" in note_start_times:
                Snare.extend(note_start_times["D2"])
            if "D1" in note_start_times:
                Hihat.extend(note_start_times["D1"])
            if "A♯0" in note_start_times:
                Hihat.extend(note_start_times["A♯0"])
            if "F♯2" in note_start_times:
                Hihat.extend(note_start_times["F♯2"])
            if "E2" in note_start_times:
                Snare.extend(note_start_times["E2"])

            # 全部含めて一番早いノードを基準にする
            min_val=min(Kick,Crash,Hihat,Snare)
            # print(min_val)

            if min_val==Kick:
                min_node_time=Kick[0]
                # print("Kick")
                # print(min_node_time)
            elif min_val==Crash:
                min_node_time=Crash[0]
                # print("Crash")
                # print(min_node_time)
            elif min_val==Hihat:
                min_node_time=Hihat[0]
                # print("Hihat")
                # print(min_node_time)
            elif min_val==Snare:
                min_node_time=Snare[0]
                # print("Snare")
                # print(min_node_time)

            Kick=sorted([n-min_node_time for n in Kick])
            # print(Kick)
            Crash=sorted([n-min_node_time for n in Crash])
            Hihat=sorted([n-min_node_time for n in Hihat])
            Snare=sorted([n-min_node_time for n in Snare])

            # 遅延処理
            # print(Kick,"\n")
            Kick_def=[]
            Crash_def=[]
            Hihat_def=[]
            Snare_def=[]
            
            Kick_tmp=copy.deepcopy(Kick)
            Crash_tmp=copy.deepcopy(Crash)
            Hihat_tmp=copy.deepcopy(Hihat)
            Snare_tmp=copy.deepcopy(Snare)
            
            # 遅延分を引く
            Kick_delayed=[n-delay for n in Kick_tmp]
            Crash_delayed=[n-delay for n in Crash_tmp]
            Hihat_delayed=[n-delay for n in Hihat_tmp]
            Snare_delayed=[n-delay for n in Snare_tmp]
            
            # ノードの時間に一番近い八分音符のバーを検索
            # そのバートの差を求める
            for l in Kick_delayed:
                tmp=getNearestValue(bar_sec,l) #return 0 は None になる
                if tmp==None: #基準のノードだったら0
                    tmp=0
                Kick_def.append(tmp-l)
            for l in Crash_delayed:
                tmp=getNearestValue(bar_sec,l)
                if tmp==None:
                    tmp=0
                Crash_def.append(tmp-l)
            for l in Hihat_delayed:
                tmp=getNearestValue(bar_sec,l)
                if tmp==None:
                    tmp=0
                Hihat_def.append(tmp-l)
            for l in Snare_delayed:
                tmp=getNearestValue(bar_sec,l)
                if tmp==None:
                    tmp=0
                Snare_def.append(tmp-l)
                    

            
            #excel書き込み
            ws["A1"].value="Kick"
            write_list_2d(ws,Kick,2,1)
            ws["B1"].value="Hihat"
            write_list_2d(ws,Hihat,2,2)
            ws["C1"].value="Snare"
            write_list_2d(ws,Snare,2,3)
            ws["D1"].value="Crash"
            write_list_2d(ws,Crash,2,4)
            
            ws["F1"].value="Kick_def"
            write_list_2d(ws,Kick_def,2,6)
            ws["G1"].value="Hihat_def"
            write_list_2d(ws,Hihat_def,2,7)
            ws["H1"].value="Snare_def"
            write_list_2d(ws,Snare_def,2,8)
            ws["I1"].value="Crash_def"
            write_list_2d(ws,Crash_def,2,9)
            
            ws["K1"].value="Kick_delayed"
            write_list_2d(ws,Kick_delayed,2,11)
            ws["L1"].value="Hihat_delayed"
            write_list_2d(ws,Hihat_delayed,2,12)
            ws["M1"].value="Snare_delayed"
            write_list_2d(ws,Snare_delayed,2,13)
            ws["N1"].value="Crash_delayed"
            write_list_2d(ws,Crash_delayed,2,14)
            
            wb.save(dir_save)
            #messageBox
            self.mb=messagebox.showinfo("通知","生成完了")
            
            
        except Exception as e:
            print(traceback.format_exc())
            title = e.__class__.__name__
            message = traceback.format_exc()
            messagebox.showerror(f"{title}",
                                 f"{message}")